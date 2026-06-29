"""
EventLedger AI – Export Engine
Generates PDF reports (ReportLab) and Excel workbooks (openpyxl).
"""

import io
import datetime

# ── PDF Export ─────────────────────────────────────────────────────────────

def generate_pdf(event, summary, income_rows, expense_rows, sponsor_rows, vendor_rows):
    """Return a bytes object containing a complete PDF financial report."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=18*mm, bottomMargin=18*mm
    )

    # ── Colors ──
    C_PRIMARY  = colors.HexColor("#6366f1")
    C_SUCCESS  = colors.HexColor("#10b981")
    C_DANGER   = colors.HexColor("#ef4444")
    C_ACCENT   = colors.HexColor("#f59e0b")
    C_BG       = colors.HexColor("#0f0f13")
    C_SURFACE  = colors.HexColor("#18181f")
    C_SURFACE2 = colors.HexColor("#22222e")
    C_TEXT     = colors.HexColor("#e8e8f0")
    C_MUTED    = colors.HexColor("#7b7b9a")
    C_WHITE    = colors.white

    styles = getSampleStyleSheet()

    def style(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    H1      = style("H1",   fontSize=22, textColor=C_TEXT,    spaceAfter=4,  fontName="Helvetica-Bold")
    H2      = style("H2",   fontSize=13, textColor=C_PRIMARY,  spaceAfter=3,  fontName="Helvetica-Bold")
    SUBTEXT = style("SUB",  fontSize=9,  textColor=C_MUTED,   spaceAfter=2)
    BODY    = style("BODY", fontSize=9,  textColor=C_TEXT,    spaceAfter=4)
    RIGHT   = style("RIGHT",fontSize=9,  textColor=C_TEXT,    alignment=TA_RIGHT)

    sym = {"USD":"$","INR":"₹","EUR":"€","GBP":"£","AED":"د.إ","SGD":"S$"}.get(event.get("currency","USD"),"$")

    def money(v):
        return f"{sym}{v:,.2f}"

    def section_table_style(header_color=C_PRIMARY):
        return TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), header_color),
            ("TEXTCOLOR",     (0,0), (-1,0), C_WHITE),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,0), 8),
            ("BOTTOMPADDING", (0,0), (-1,0), 6),
            ("TOPPADDING",    (0,0), (-1,0), 6),
            ("BACKGROUND",    (0,1), (-1,-1), C_SURFACE),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [C_SURFACE, C_SURFACE2]),
            ("TEXTCOLOR",     (0,1), (-1,-1), C_TEXT),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",      (0,1), (-1,-1), 8),
            ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#2e2e3e")),
            ("TOPPADDING",    (0,1), (-1,-1), 4),
            ("BOTTOMPADDING", (0,1), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        ])

    story = []

    # ── Cover header ──
    cover = Table([[
        Paragraph(f"<b>EventLedger AI</b>", style("BRD", fontSize=10, textColor=C_PRIMARY, fontName="Helvetica-Bold")),
        Paragraph(f"Financial Report &nbsp;·&nbsp; {datetime.date.today()}", style("DT", fontSize=9, textColor=C_MUTED, alignment=TA_RIGHT)),
    ]], colWidths=["60%","40%"])
    cover.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,-1), C_BG),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ("TOPPADDING",  (0,0),(-1,-1), 8),
        ("LEFTPADDING", (0,0),(-1,-1), 0),
        ("RIGHTPADDING",(0,0),(-1,-1), 0),
    ]))
    story.append(cover)
    story.append(HRFlowable(width="100%", thickness=1, color=C_PRIMARY, spaceAfter=12))

    # ── Title block ──
    story.append(Paragraph(event.get("name","Event"), H1))
    story.append(Paragraph(
        f"Venue: {event.get('venue','—')} &nbsp;|&nbsp; "
        f"{event.get('start_date','—')} → {event.get('end_date','—')} &nbsp;|&nbsp; "
        f"{event.get('expected_attendees',0):,} attendees",
        SUBTEXT
    ))
    story.append(Spacer(1, 8*mm))

    # ── KPI Summary row ──
    profit = summary["act_profit"]
    margin = (profit / max(summary["act_income"],1)) * 100
    kpi_data = [
        ["Total Revenue", "Total Expenses", "Net Profit", "Margin", "Budget Accuracy"],
        [money(summary["act_income"]), money(summary["act_expense"]),
         money(profit), f"{margin:.1f}%", f"{summary['budget_accuracy']:.1f}%"],
    ]
    kpi_tbl = Table(kpi_data, colWidths=["20%","20%","20%","20%","20%"])
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), C_SURFACE2),
        ("TEXTCOLOR",     (0,0), (-1,0), C_MUTED),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica"),
        ("FONTSIZE",      (0,0), (-1,0), 7),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("BACKGROUND",    (0,1), (-1,1), C_SURFACE),
        ("TEXTCOLOR",     (0,1), (-1,1), C_TEXT),
        ("FONTNAME",      (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",      (0,1), (-1,1), 11),
        ("TEXTCOLOR",     (2,1), (2,1), C_SUCCESS if profit >= 0 else C_DANGER),
        ("TEXTCOLOR",     (3,1), (3,1), C_SUCCESS if margin >= 0 else C_DANGER),
        ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#2e2e3e")),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("ROUNDEDCORNERS",(0,0), (-1,-1), 4),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 8*mm))

    # ── P&L Table ──
    story.append(Paragraph("Profit & Loss Statement", H2))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_PRIMARY, spaceAfter=6))
    pl_data = [
        ["Line Item", "Estimated", "Actual", "Variance"],
        ["Total Income",
         money(summary["est_income"]), money(summary["act_income"]),
         money(summary["income_variance"])],
        ["Total Expenses",
         money(summary["est_expense"]), money(summary["act_expense"]),
         money(summary["expense_variance"])],
        ["Net Profit",
         money(summary["est_profit"]), money(summary["act_profit"]),
         money(summary["act_profit"] - summary["est_profit"])],
        ["Budget Accuracy", "—", "—", f"{summary['budget_accuracy']:.1f}%"],
    ]
    pl_tbl = Table(pl_data, colWidths=["40%","20%","20%","20%"])
    pl_tbl.setStyle(section_table_style(C_PRIMARY))
    story.append(pl_tbl)
    story.append(Spacer(1, 8*mm))

    # ── Income Table ──
    if income_rows:
        story.append(Paragraph("Income Details", H2))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C_SUCCESS, spaceAfter=6))
        inc_data = [["Source", "Category", "Amount", "Date", "Mode"]]
        for r in income_rows:
            inc_data.append([
                r.get("source",""),
                r.get("category",""),
                money(r.get("amount",0)),
                r.get("received_on","—"),
                r.get("payment_mode","—"),
            ])
        total_inc = sum(r.get("amount",0) for r in income_rows)
        inc_data.append(["TOTAL", "", money(total_inc), "", ""])
        inc_tbl = Table(inc_data, colWidths=["30%","20%","18%","18%","14%"])
        style2 = section_table_style(C_SUCCESS)
        style2.add("FONTNAME",  (0,-1), (-1,-1), "Helvetica-Bold")
        style2.add("BACKGROUND",(0,-1), (-1,-1), C_SURFACE2)
        style2.add("TEXTCOLOR", (2,-1), (2,-1), C_SUCCESS)
        inc_tbl.setStyle(style2)
        story.append(inc_tbl)
        story.append(Spacer(1, 8*mm))

    # ── Expense Table ──
    if expense_rows:
        story.append(Paragraph("Expense Details", H2))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C_DANGER, spaceAfter=6))
        exp_data = [["Department", "Category", "Description", "Amount", "Date"]]
        for r in expense_rows:
            exp_data.append([
                r.get("dept_name","—"),
                r.get("category",""),
                (r.get("description","") or "")[:30],
                money(r.get("amount",0)),
                r.get("paid_on","—"),
            ])
        total_exp = sum(r.get("amount",0) for r in expense_rows)
        exp_data.append(["TOTAL", "", "", money(total_exp), ""])
        exp_tbl = Table(exp_data, colWidths=["22%","18%","26%","18%","16%"])
        style3 = section_table_style(C_DANGER)
        style3.add("FONTNAME",  (0,-1), (-1,-1), "Helvetica-Bold")
        style3.add("BACKGROUND",(0,-1), (-1,-1), C_SURFACE2)
        style3.add("TEXTCOLOR", (3,-1), (3,-1), C_DANGER)
        exp_tbl.setStyle(style3)
        story.append(exp_tbl)
        story.append(Spacer(1, 8*mm))

    # ── Sponsors Table ──
    if sponsor_rows:
        story.append(Paragraph("Sponsors", H2))
        story.append(HRFlowable(width="100%", thickness=0.5, color=C_ACCENT, spaceAfter=6))
        sp_data = [["Sponsor", "Tier", "Contact", "Amount", "Status"]]
        for s in sponsor_rows:
            sp_data.append([
                s.get("name",""),
                s.get("tier",""),
                s.get("contact_name","—"),
                money(s.get("amount",0)),
                s.get("status","confirmed"),
            ])
        sp_tbl = Table(sp_data, colWidths=["28%","16%","22%","18%","16%"])
        sp_tbl.setStyle(section_table_style(C_ACCENT))
        story.append(sp_tbl)
        story.append(Spacer(1, 8*mm))

    # ── Vendors Table ──
    if vendor_rows:
        story.append(Paragraph("Vendors", H2))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#8b5cf6"), spaceAfter=6))
        vd_data = [["Vendor", "Category", "Contact", "Contract Value", "Status"]]
        for v in vendor_rows:
            vd_data.append([
                v.get("name",""),
                v.get("category",""),
                v.get("contact_name","—"),
                money(v.get("contract_value",0)),
                v.get("status","active"),
            ])
        vd_tbl = Table(vd_data, colWidths=["26%","20%","22%","18%","14%"])
        vd_tbl.setStyle(section_table_style(colors.HexColor("#8b5cf6")))
        story.append(vd_tbl)

    # ── Footer ──
    story.append(Spacer(1, 10*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_SURFACE2, spaceAfter=4))
    story.append(Paragraph(
        f"Generated by EventLedger AI &nbsp;·&nbsp; {datetime.datetime.now().strftime('%d %b %Y %H:%M')}",
        style("FOOT", fontSize=7, textColor=C_MUTED, alignment=TA_CENTER)
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ── Excel Export ───────────────────────────────────────────────────────────

def generate_excel(event, summary, income_rows, expense_rows, sponsor_rows, vendor_rows, dept_rows):
    """Return bytes of an Excel workbook with multiple sheets."""
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    sym = {"USD":"$","INR":"₹","EUR":"€","GBP":"£","AED":"AED ","SGD":"S$"}.get(event.get("currency","USD"),"$")

    # ── Colors ──
    BG       = "0F0F13"
    SURFACE  = "18181F"
    SURFACE2 = "22222E"
    PRIMARY  = "6366F1"
    SUCCESS  = "10B981"
    DANGER   = "EF4444"
    ACCENT   = "F59E0B"
    TEXT     = "E8E8F0"
    MUTED    = "7B7B9A"
    WHITE    = "FFFFFF"

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def cell_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    thin = Side(style="thin", color="2E2E3E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell, color=PRIMARY):
        cell.fill = hdr_fill(color)
        cell.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    def style_data(cell, bold=False, color=TEXT, fill=SURFACE, align="left"):
        cell.fill = cell_fill(fill)
        cell.font = Font(bold=bold, color=color, size=9, name="Calibri")
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = border

    def style_total(cell, color=WHITE):
        cell.fill = hdr_fill(SURFACE2)
        cell.font = Font(bold=True, color=color, size=9, name="Calibri")
        cell.alignment = Alignment(horizontal="right")
        cell.border = border

    def auto_width(ws, padding=4):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + padding, 45)

    def set_tab_color(ws, hex_color):
        ws.sheet_properties.tabColor = hex_color

    # ════════════════════════════════════
    # Sheet 1 – Summary Dashboard
    # ════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    set_tab_color(ws, PRIMARY)
    ws.sheet_view.showGridLines = False

    # Title rows
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"EventLedger AI — {event.get('name','')} Financial Report"
    c.font   = Font(bold=True, color=PRIMARY, size=14, name="Calibri")
    c.fill   = cell_fill(SURFACE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    c2 = ws["A2"]
    c2.value = (f"Venue: {event.get('venue','—')}  |  "
                f"{event.get('start_date','—')} → {event.get('end_date','—')}  |  "
                f"{event.get('expected_attendees',0):,} attendees  |  "
                f"Generated: {datetime.date.today()}")
    c2.font  = Font(color=MUTED, size=9, name="Calibri")
    c2.fill  = cell_fill(SURFACE)
    ws.row_dimensions[2].height = 16

    ws.append([])

    # KPI block
    kpi_headers = ["Metric", "Estimated", "Actual", "Variance"]
    ws.append(kpi_headers)
    for col_idx, h in enumerate(kpi_headers, 1):
        style_header(ws.cell(ws.max_row, col_idx), PRIMARY)

    kpis = [
        ("Total Income",    summary["est_income"],   summary["act_income"],   summary["income_variance"]),
        ("Total Expenses",  summary["est_expense"],  summary["act_expense"],  summary["expense_variance"]),
        ("Net Profit",      summary["est_profit"],   summary["act_profit"],   summary["act_profit"]-summary["est_profit"]),
        ("Budget Accuracy", f"{summary['budget_accuracy']:.1f}%", "—",       "—"),
    ]
    for i, (label, est, act, var) in enumerate(kpis):
        ws.append([label, est, act, var])
        row = ws.max_row
        fill = SURFACE if i % 2 == 0 else SURFACE2
        for col_idx in range(1, 5):
            cell = ws.cell(row, col_idx)
            color = TEXT
            if col_idx == 4 and isinstance(var, (int, float)):
                color = SUCCESS if var >= 0 else DANGER
            style_data(cell, fill=fill, color=color, align="right" if col_idx > 1 else "left")
        ws.row_dimensions[row].height = 18

    auto_width(ws)

    # ════════════════════════════════════
    # Sheet 2 – Income
    # ════════════════════════════════════
    ws2 = wb.create_sheet("Income")
    set_tab_color(ws2, SUCCESS)
    ws2.sheet_view.showGridLines = False
    headers = ["Source", "Category", "Amount", "Date", "Mode", "Reference", "Notes"]
    ws2.append(headers)
    for ci, h in enumerate(headers, 1):
        style_header(ws2.cell(1, ci), SUCCESS)
    total = 0
    for i, r in enumerate(income_rows):
        amt = r.get("amount", 0)
        total += amt
        ws2.append([r.get("source",""), r.get("category",""), amt,
                    r.get("received_on",""), r.get("payment_mode",""),
                    r.get("reference",""), r.get("notes","")])
        row = ws2.max_row
        fill = SURFACE if i%2==0 else SURFACE2
        for ci in range(1, 8):
            style_data(ws2.cell(row, ci), fill=fill,
                       align="right" if ci==3 else "left",
                       color=SUCCESS if ci==3 else TEXT)
        ws2.row_dimensions[row].height = 16
    ws2.append(["TOTAL", "", total, "", "", "", ""])
    trow = ws2.max_row
    for ci in range(1, 8):
        style_total(ws2.cell(trow, ci), color=SUCCESS if ci==3 else TEXT)
    auto_width(ws2)

    # ════════════════════════════════════
    # Sheet 3 – Expenses
    # ════════════════════════════════════
    ws3 = wb.create_sheet("Expenses")
    set_tab_color(ws3, DANGER)
    ws3.sheet_view.showGridLines = False
    headers3 = ["Department", "Category", "Description", "Amount", "Date", "Mode", "Status"]
    ws3.append(headers3)
    for ci, h in enumerate(headers3, 1):
        style_header(ws3.cell(1, ci), DANGER)
    total3 = 0
    for i, r in enumerate(expense_rows):
        amt = r.get("amount", 0)
        total3 += amt
        ws3.append([r.get("dept_name","—"), r.get("category",""),
                    r.get("description",""), amt,
                    r.get("paid_on",""), r.get("payment_mode",""), r.get("status","")])
        row = ws3.max_row
        fill = SURFACE if i%2==0 else SURFACE2
        for ci in range(1, 8):
            style_data(ws3.cell(row, ci), fill=fill,
                       align="right" if ci==4 else "left",
                       color=DANGER if ci==4 else TEXT)
        ws3.row_dimensions[row].height = 16
    ws3.append(["TOTAL", "", "", total3, "", "", ""])
    trow3 = ws3.max_row
    for ci in range(1, 8):
        style_total(ws3.cell(trow3, ci), color=DANGER if ci==4 else TEXT)
    auto_width(ws3)

    # ════════════════════════════════════
    # Sheet 4 – Sponsors
    # ════════════════════════════════════
    ws4 = wb.create_sheet("Sponsors")
    set_tab_color(ws4, ACCENT)
    ws4.sheet_view.showGridLines = False
    headers4 = ["Sponsor", "Tier", "Contact", "Email", "Amount", "Status", "Notes"]
    ws4.append(headers4)
    for ci, h in enumerate(headers4, 1):
        style_header(ws4.cell(1, ci), ACCENT)
    total4 = 0
    tier_colors = {"Platinum":"D1D5DB","Gold":"FCD34D","Silver":"9CA3AF","Bronze":"B45309"}
    for i, s in enumerate(sponsor_rows):
        amt = s.get("amount", 0)
        total4 += amt
        ws4.append([s.get("name",""), s.get("tier",""),
                    s.get("contact_name",""), s.get("contact_email",""),
                    amt, s.get("status","confirmed"), s.get("notes","")])
        row = ws4.max_row
        fill = SURFACE if i%2==0 else SURFACE2
        for ci in range(1, 8):
            tc = tier_colors.get(s.get("tier",""), TEXT) if ci==2 else (SUCCESS if ci==5 else TEXT)
            style_data(ws4.cell(row, ci), fill=fill,
                       align="right" if ci==5 else "left", color=tc)
        ws4.row_dimensions[row].height = 16
    if sponsor_rows:
        ws4.append(["TOTAL", "", "", "", total4, "", ""])
        trow4 = ws4.max_row
        for ci in range(1, 8):
            style_total(ws4.cell(trow4, ci), color=SUCCESS if ci==5 else TEXT)
    auto_width(ws4)

    # ════════════════════════════════════
    # Sheet 5 – Vendors
    # ════════════════════════════════════
    ws5 = wb.create_sheet("Vendors")
    set_tab_color(ws5, "8B5CF6")
    ws5.sheet_view.showGridLines = False
    headers5 = ["Vendor", "Category", "Contact", "Email", "Contract Value", "Status", "Notes"]
    ws5.append(headers5)
    for ci, h in enumerate(headers5, 1):
        style_header(ws5.cell(1, ci), "8B5CF6")
    total5 = 0
    for i, v in enumerate(vendor_rows):
        cv = v.get("contract_value", 0)
        total5 += cv
        ws5.append([v.get("name",""), v.get("category",""),
                    v.get("contact_name",""), v.get("contact_email",""),
                    cv, v.get("status","active"), v.get("notes","")])
        row = ws5.max_row
        fill = SURFACE if i%2==0 else SURFACE2
        for ci in range(1, 8):
            style_data(ws5.cell(row, ci), fill=fill,
                       align="right" if ci==5 else "left",
                       color=ACCENT if ci==5 else TEXT)
        ws5.row_dimensions[row].height = 16
    if vendor_rows:
        ws5.append(["TOTAL", "", "", "", total5, "", ""])
        trow5 = ws5.max_row
        for ci in range(1, 8):
            style_total(ws5.cell(trow5, ci), color=ACCENT if ci==5 else TEXT)
    auto_width(ws5)

    # ════════════════════════════════════
    # Sheet 6 – Departments
    # ════════════════════════════════════
    if dept_rows:
        ws6 = wb.create_sheet("Departments")
        set_tab_color(ws6, "06B6D4")
        ws6.sheet_view.showGridLines = False
        headers6 = ["Department", "Est. Expense", "Act. Expense", "Variance", "% Variance", "Status"]
        ws6.append(headers6)
        for ci, h in enumerate(headers6, 1):
            style_header(ws6.cell(1, ci), "06B6D4")
        for i, d in enumerate(dept_rows):
            var = d["act"] - d["est"]
            pct = (var / max(d["est"],1)) * 100
            status = "Over Budget" if var > 0 else "Under Budget" if var < 0 else "On Track"
            ws6.append([d["name"], d["est"], d["act"], var, f"{pct:.1f}%", status])
            row = ws6.max_row
            fill = SURFACE if i%2==0 else SURFACE2
            for ci in range(1, 7):
                col = (DANGER if var > 0 else SUCCESS) if ci in (4,5,6) else TEXT
                style_data(ws6.cell(row, ci), fill=fill,
                           align="right" if ci>1 else "left", color=col)
            ws6.row_dimensions[row].height = 16
        auto_width(ws6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
