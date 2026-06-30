"""
EventLedger AI – Excel Import Engine
Generates a downloadable template (Income + Expenses sheets) and
imports a filled-in template into a SPECIFIC event, without touching
any other event's data.
"""

import io
import datetime
import pandas as pd

INCOME_COLUMNS = [
    "Source", "Category", "Amount", "Received On (YYYY-MM-DD)",
    "Payment Mode", "Reference", "Notes"
]

EXPENSE_COLUMNS = [
    "Department", "Category", "Item Name", "Description", "Quantity",
    "Unit", "Amount", "Paid On (YYYY-MM-DD)", "Payment Mode", "Status",
    "Reference", "Notes"
]

PAY_MODES = ["Cash", "Bank Transfer", "UPI", "Card", "Cheque", "Online", "DD"]
STATUSES  = ["paid", "pending", "partial"]


def generate_template(dept_names: list) -> bytes:
    """
    Returns an .xlsx template with Income and Expenses sheets,
    pre-filled with one example row, plus a Departments reference sheet
    so the user knows valid department names to type in.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="0F0F13")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    EXAMPLE_FONT = Font(italic=True, color="888888", size=9, name="Calibri")

    def write_sheet(ws, columns, example_row):
        for i, col in enumerate(columns, start=1):
            c = ws.cell(row=1, column=i, value=col)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center")
        for i, val in enumerate(example_row, start=1):
            c = ws.cell(row=2, column=i, value=val)
            c.font = EXAMPLE_FONT
        for i, col in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(len(col) + 2, 16)

    ws_inc = wb.active
    ws_inc.title = "Income"
    write_sheet(ws_inc, INCOME_COLUMNS, [
        "Ticket Sales", "Tickets", 50000, str(datetime.date.today()),
        "Bank Transfer", "TXN12345", "Example row — delete before importing"
    ])

    ws_exp = wb.create_sheet("Expenses")
    write_sheet(ws_exp, EXPENSE_COLUMNS, [
        dept_names[0] if dept_names else "Marketing", "Venue", "Hall Booking",
        "Main auditorium, 2 days", 1, "unit", 25000, str(datetime.date.today()),
        "Bank Transfer", "paid", "INV-001", "Example row — delete before importing"
    ])

    ws_ref = wb.create_sheet("Valid Departments")
    ws_ref.cell(row=1, column=1, value="Department Name").font = HEADER_FONT
    ws_ref.cell(row=1, column=1).fill = HEADER_FILL
    for i, d in enumerate(dept_names, start=2):
        ws_ref.cell(row=i, column=1, value=d)
    ws_ref.column_dimensions["A"].width = 30
    if not dept_names:
        ws_ref.cell(row=2, column=1, value="(No departments yet — add one in Planning first)")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_and_validate(file_bytes: bytes, dept_name_to_id: dict) -> dict:
    """
    Reads the uploaded Excel file and validates every row.
    Returns:
      {
        "ok": bool,
        "income_rows": [...],   # cleaned dicts ready to insert
        "expense_rows": [...],
        "errors": [str, ...],   # row-level problems, doesn't block valid rows
        "message": str
      }
    """
    errors = []
    income_rows = []
    expense_rows = []

    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
    except Exception as e:
        return {"ok": False, "income_rows": [], "expense_rows": [],
                "errors": [], "message": f"Could not read Excel file: {e}"}

    # ── Income sheet ──────────────────────────────────────────────────────
    if "Income" in xls.sheet_names:
        df = xls.parse("Income")
        for idx, row in df.iterrows():
            source = str(row.get("Source", "")).strip()
            if not source or source.lower() == "nan" or "example row" in str(row.get("Notes","")).lower():
                continue
            try:
                amount = float(row.get("Amount", 0) or 0)
            except Exception:
                errors.append(f"Income row {idx+2}: invalid amount, skipped")
                continue
            if amount <= 0:
                errors.append(f"Income row {idx+2}: amount must be greater than 0, skipped")
                continue
            received_on = str(row.get("Received On (YYYY-MM-DD)", "") or datetime.date.today())
            mode = str(row.get("Payment Mode", "") or "Cash").strip()
            if mode not in PAY_MODES:
                mode = "Cash"
            income_rows.append({
                "source": source,
                "category": str(row.get("Category", "Other") or "Other").strip(),
                "amount": amount,
                "received_on": received_on,
                "payment_mode": mode,
                "reference": str(row.get("Reference", "") or "").strip(),
                "notes": str(row.get("Notes", "") or "").strip(),
            })

    # ── Expenses sheet ────────────────────────────────────────────────────
    if "Expenses" in xls.sheet_names:
        df = xls.parse("Expenses")
        for idx, row in df.iterrows():
            item_name = str(row.get("Item Name", "")).strip()
            if not item_name or item_name.lower() == "nan" or "example row" in str(row.get("Notes","")).lower():
                continue
            dept_name = str(row.get("Department", "")).strip()
            dept_id = dept_name_to_id.get(dept_name)
            if not dept_id:
                errors.append(f"Expense row {idx+2}: unknown department '{dept_name}', skipped")
                continue
            try:
                amount = float(row.get("Amount", 0) or 0)
            except Exception:
                errors.append(f"Expense row {idx+2}: invalid amount, skipped")
                continue
            if amount <= 0:
                errors.append(f"Expense row {idx+2}: amount must be greater than 0, skipped")
                continue
            try:
                qty = float(row.get("Quantity", 1) or 1)
            except Exception:
                qty = 1
            paid_on = str(row.get("Paid On (YYYY-MM-DD)", "") or datetime.date.today())
            mode = str(row.get("Payment Mode", "") or "Cash").strip()
            if mode not in PAY_MODES:
                mode = "Cash"
            status = str(row.get("Status", "") or "paid").strip().lower()
            if status not in STATUSES:
                status = "paid"
            expense_rows.append({
                "dept_id": dept_id,
                "dept_name": dept_name,
                "category": str(row.get("Category", "Miscellaneous") or "Miscellaneous").strip(),
                "item_name": item_name,
                "description": str(row.get("Description", "") or "").strip(),
                "quantity": qty,
                "unit": str(row.get("Unit", "unit") or "unit").strip(),
                "amount": amount,
                "paid_on": paid_on,
                "payment_mode": mode,
                "status": status,
                "reference": str(row.get("Reference", "") or "").strip(),
                "notes": str(row.get("Notes", "") or "").strip(),
            })

    if not income_rows and not expense_rows:
        return {"ok": False, "income_rows": [], "expense_rows": [], "errors": errors,
                "message": "No valid rows found. Check the template format and try again."}

    return {"ok": True, "income_rows": income_rows, "expense_rows": expense_rows,
            "errors": errors, "message": "File parsed successfully"}


def import_to_event(event_id: int, income_rows: list, expense_rows: list) -> dict:
    """
    Inserts the parsed rows into the database for a SPECIFIC event only.
    Uses the existing helper functions so all normal validation/columns apply.
    """
    from utils.helpers import add_actual_income, add_actual_expense

    inc_count = 0
    exp_count = 0

    for r in income_rows:
        add_actual_income(
            event_id, r["source"], r["category"], r["amount"],
            r["received_on"], r["payment_mode"], r["reference"], r["notes"]
        )
        inc_count += 1

    for r in expense_rows:
        add_actual_expense(
            event_id, r["dept_id"], r["category"], r["item_name"], r["description"],
            r["amount"], r["paid_on"], r["payment_mode"], r["status"],
            quantity=r["quantity"], unit=r["unit"], reference=r["reference"], notes=r["notes"]
        )
        exp_count += 1

    return {"ok": True, "income_imported": inc_count, "expense_imported": exp_count}
